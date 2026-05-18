#!/usr/bin/env python3
"""Evaluate Gemini STT (default: gemini-2.5-flash) on wav/txt pairs.

Each row: send audio to Gemini, compare transcript to matching .txt ground truth,
report WER, CER, token usage, and latency. Writes per-row XLSX + summary JSON.

Requires GEMINI_API_KEY in the environment.

Example (remote PC folder from screenshot):
  set GEMINI_API_KEY=your_key
  python scripts/evaluate_gemini_stt.py ^
    --data-dir "E:\\STT DATA\\TRAINING_PROCESSED\\RadioTavisufleba_ForTraining" ^
    --sample-size 50 --seed 42

Pilot (no API calls, estimates only):
  python scripts/evaluate_gemini_stt.py --data-dir "..." --dry-run --sample-size 50
"""

from __future__ import annotations

import argparse
import json
import os
import random
import statistics
import sys
import time
from dataclasses import asdict, dataclass
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency: openpyxl. Install with: pip install openpyxl"
    ) from exc

def _require_genai():
    try:
        from google import genai
        from google.genai import types
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise SystemExit(
            "Missing dependency: google-genai. Install with: pip install google-genai"
        ) from exc
    return genai, types


DEFAULT_MODEL = "gemini-2.5-flash"
TRANSCRIBE_PROMPT = (
    "Transcribe the spoken Georgian in this audio clip. "
    "Return only the transcript text in Georgian. "
    "No timestamps, labels, translation, or explanation."
)

# Gemini audio input is roughly 32 tokens per second (API docs).
AUDIO_TOKENS_PER_SECOND = 32
PROMPT_TOKENS_ESTIMATE = 35
OUTPUT_TOKENS_ESTIMATE = 40


@dataclass
class AudioPair:
    stem: str
    wav_path: Path
    txt_path: Path
    reference_text: str = ""


@dataclass
class RowMetrics:
    stem: str
    wav_path: str
    reference_text: str
    prediction: str
    wer: float
    cer: float
    execution_time_s: float
    prompt_tokens: int
    output_tokens: int
    total_tokens: int
    error: str = ""


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _tokenize(text: str) -> list[str]:
    return text.split()


def _levenshtein_distance(seq_a: list[str], seq_b: list[str]) -> int:
    if not seq_a:
        return len(seq_b)
    if not seq_b:
        return len(seq_a)

    prev = list(range(len(seq_b) + 1))
    for i, a in enumerate(seq_a, start=1):
        curr = [i] + [0] * len(seq_b)
        for j, b in enumerate(seq_b, start=1):
            cost = 0 if a == b else 1
            curr[j] = min(
                prev[j] + 1,
                curr[j - 1] + 1,
                prev[j - 1] + cost,
            )
        prev = curr
    return prev[-1]


def _wer(reference: str, hypothesis: str) -> float:
    ref_words = _tokenize(reference)
    hyp_words = _tokenize(hypothesis)
    if not ref_words:
        return 0.0 if not hyp_words else 1.0
    return _levenshtein_distance(ref_words, hyp_words) / len(ref_words)


def _cer(reference: str, hypothesis: str) -> float:
    ref_chars = list(reference)
    hyp_chars = list(hypothesis)
    if not ref_chars:
        return 0.0 if not hyp_chars else 1.0
    return _levenshtein_distance(ref_chars, hyp_chars) / len(ref_chars)


def _discover_pairs(data_dir: Path) -> list[AudioPair]:
    if not data_dir.is_dir():
        raise SystemExit(f"Data directory not found: {data_dir}")

    pairs: list[AudioPair] = []
    for wav_path in sorted(data_dir.glob("*.wav")):
        txt_path = wav_path.with_suffix(".txt")
        if not txt_path.is_file():
            continue
        reference = _normalize_text(txt_path.read_text(encoding="utf-8"))
        pairs.append(
            AudioPair(
                stem=wav_path.stem,
                wav_path=wav_path,
                txt_path=txt_path,
                reference_text=reference,
            )
        )
    return pairs


def _select_pairs(
    pairs: list[AudioPair],
    sample_size: int | None,
    seed: int,
    max_rows: int | None,
) -> list[AudioPair]:
    selected = list(pairs)
    if sample_size is not None and sample_size < len(selected):
        rng = random.Random(seed)
        selected = rng.sample(selected, sample_size)
        selected.sort(key=lambda p: p.stem)
    if max_rows is not None:
        selected = selected[:max_rows]
    return selected


def _estimate_audio_seconds(wav_path: Path) -> float:
    """Rough duration from WAV header (PCM); fallback to size heuristic."""
    try:
        import wave

        with wave.open(str(wav_path), "rb") as wf:
            rate = wf.getframerate() or 16000
            frames = wf.getnframes()
            if rate > 0 and frames > 0:
                return frames / rate
    except (OSError, wave.Error):
        pass
    # ~16 kHz mono 16-bit ≈ 32 KB/s
    size_kb = wav_path.stat().st_size / 1024
    return max(1.0, size_kb / 32.0)


def _estimate_tokens_for_pair(pair: AudioPair) -> tuple[int, int, int]:
    seconds = _estimate_audio_seconds(pair.wav_path)
    prompt_tokens = PROMPT_TOKENS_ESTIMATE + int(seconds * AUDIO_TOKENS_PER_SECOND)
    output_tokens = max(
        OUTPUT_TOKENS_ESTIMATE,
        len(_tokenize(pair.reference_text)) + 5,
    )
    return prompt_tokens, output_tokens, prompt_tokens + output_tokens


def _print_quota_guidance(total_pairs: int, est_total_tokens: int) -> None:
    print(
        "\n--- free-tier planning (check your limits in AI Studio) ---\n"
        f"Pairs selected: {total_pairs}\n"
        f"Estimated total tokens (all rows): ~{est_total_tokens:,}\n"
        "Typical free-tier limits for gemini-2.5-flash (vary by account):\n"
        "  • ~10 requests/minute  -> use --delay-s 6.5 or higher\n"
        "  • ~250-1,500 requests/day -> 500 clips may need 2 days or sampling\n"
        "  • ~250,000 input tokens/minute\n"
        "\nRecommended sample sizes for WER/CER (Georgian STT):\n"
        "  • 10 rows  - smoke test / API wiring\n"
        "  • 50 rows  - quick estimate (often +/-5% on mean WER)\n"
        "  • 100 rows - solid benchmark for comparing models\n"
        "  • 200 rows - high-confidence aggregate metrics\n"
        "  • 500 rows - full set; only if daily request quota allows\n"
    )


def _load_checkpoint(path: Path) -> dict[str, dict]:
    if not path.is_file():
        return {}
    raw = json.loads(path.read_text(encoding="utf-8"))
    rows = raw.get("rows", raw)
    if not isinstance(rows, dict):
        return {}
    return rows


def _save_checkpoint(path: Path, rows: dict[str, dict]) -> None:
    payload = {"rows": rows}
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _usage_counts(response: object) -> tuple[int, int, int]:
    usage = getattr(response, "usage_metadata", None)
    if usage is None:
        return 0, 0, 0
    prompt = int(getattr(usage, "prompt_token_count", 0) or 0)
    output = int(getattr(usage, "candidates_token_count", 0) or 0)
    total = int(getattr(usage, "total_token_count", 0) or 0)
    if total == 0:
        total = prompt + output
    return prompt, output, total


def _transcribe(
    client: object,
    model: str,
    wav_path: Path,
    temperature: float,
) -> tuple[str, int, int, int]:
    _, types = _require_genai()
    audio_bytes = wav_path.read_bytes()
    response = client.models.generate_content(
        model=model,
        contents=[
            types.Content(
                role="user",
                parts=[
                    types.Part.from_bytes(data=audio_bytes, mime_type="audio/wav"),
                    types.Part.from_text(text=TRANSCRIBE_PROMPT),
                ],
            )
        ],
        config=types.GenerateContentConfig(
            temperature=temperature,
            max_output_tokens=1024,
        ),
    )
    text = _normalize_text(getattr(response, "text", "") or "")
    prompt_tokens, output_tokens, total_tokens = _usage_counts(response)
    return text, prompt_tokens, output_tokens, total_tokens


def _append_or_create_xlsx(output_xlsx: Path, metrics: RowMetrics) -> None:
    headers = [
        "stem",
        "wav_path",
        "reference_text",
        "prediction",
        "wer",
        "cer",
        "execution_time_s",
        "prompt_tokens",
        "output_tokens",
        "total_tokens",
        "error",
    ]
    row = [
        metrics.stem,
        metrics.wav_path,
        metrics.reference_text,
        metrics.prediction,
        metrics.wer,
        metrics.cer,
        metrics.execution_time_s,
        metrics.prompt_tokens,
        metrics.output_tokens,
        metrics.total_tokens,
        metrics.error,
    ]

    if output_xlsx.is_file():
        wb = load_workbook(output_xlsx)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "gemini_stt_eval"
        ws.append(headers)
    ws.append(row)
    wb.save(output_xlsx)


def _write_summary(
    summary_json: Path,
    data_dir: Path,
    model: str,
    metrics: list[RowMetrics],
    skipped: int,
) -> dict:
    ok = [m for m in metrics if not m.error]
    if ok:
        summary = {
            "data_dir": str(data_dir),
            "model": model,
            "rows_processed": len(metrics),
            "rows_scored": len(ok),
            "rows_failed": len(metrics) - len(ok),
            "rows_skipped_resume": skipped,
            "average_wer": statistics.fmean(x.wer for x in ok),
            "median_wer": statistics.median(x.wer for x in ok),
            "average_cer": statistics.fmean(x.cer for x in ok),
            "median_cer": statistics.median(x.cer for x in ok),
            "average_execution_time_s": statistics.fmean(x.execution_time_s for x in ok),
            "total_prompt_tokens": sum(x.prompt_tokens for x in ok),
            "total_output_tokens": sum(x.output_tokens for x in ok),
            "total_tokens": sum(x.total_tokens for x in ok),
            "total_execution_time_s": sum(x.execution_time_s for x in ok),
        }
    else:
        summary = {
            "data_dir": str(data_dir),
            "model": model,
            "rows_processed": len(metrics),
            "rows_scored": 0,
            "rows_failed": len(metrics),
            "rows_skipped_resume": skipped,
        }
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def run_evaluation(args: argparse.Namespace) -> dict:
    data_dir = args.data_dir.resolve()
    pairs = _discover_pairs(data_dir)
    if not pairs:
        raise SystemExit(f"No wav/txt pairs found under {data_dir}")

    selected = _select_pairs(pairs, args.sample_size, args.seed, args.max_rows)
    estimates = [_estimate_tokens_for_pair(p) for p in selected]
    est_total_tokens = sum(x[2] for x in estimates)
    _print_quota_guidance(len(selected), est_total_tokens)

    if args.dry_run:
        per_row = []
        for pair, (prompt_t, out_t, total_t) in zip(selected, estimates):
            per_row.append(
                {
                    "stem": pair.stem,
                    "wav_kb": round(pair.wav_path.stat().st_size / 1024, 1),
                    "est_audio_seconds": round(_estimate_audio_seconds(pair.wav_path), 2),
                    "ref_words": len(_tokenize(pair.reference_text)),
                    "est_prompt_tokens": prompt_t,
                    "est_output_tokens": out_t,
                    "est_total_tokens": total_t,
                }
            )
        dry = {
            "dry_run": True,
            "data_dir": str(data_dir),
            "pairs_found": len(pairs),
            "pairs_selected": len(selected),
            "est_total_tokens": est_total_tokens,
            "est_requests": len(selected),
            "rows": per_row,
        }
        args.summary_json.write_text(
            json.dumps(dry, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(json.dumps(dry, ensure_ascii=False, indent=2))
        return dry

    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key:
        raise SystemExit(
            "Set GEMINI_API_KEY before running (get a key at https://aistudio.google.com/apikey)."
        )

    genai, _ = _require_genai()
    client = genai.Client(api_key=api_key)
    checkpoint = _load_checkpoint(args.checkpoint_json) if args.resume else {}
    metrics: list[RowMetrics] = []
    skipped = 0
    started_at = time.perf_counter()

    for idx, pair in enumerate(selected, start=1):
        if pair.stem in checkpoint:
            skipped += 1
            cached = checkpoint[pair.stem]
            metrics.append(RowMetrics(**cached))
            if args.progress_every > 0 and idx % args.progress_every == 0:
                print(f"[progress] {idx}/{len(selected)} (resume skip: {pair.stem})", flush=True)
            continue

        if idx > 1 and args.delay_s > 0:
            time.sleep(args.delay_s)

        start_t = time.perf_counter()
        error = ""
        prediction = ""
        prompt_tokens = output_tokens = total_tokens = 0
        try:
            prediction, prompt_tokens, output_tokens, total_tokens = _transcribe(
                client,
                args.model,
                pair.wav_path,
                args.temperature,
            )
        except Exception as exc:  # noqa: BLE001 — collect per-row API failures
            error = str(exc)

        elapsed = time.perf_counter() - start_t
        row = RowMetrics(
            stem=pair.stem,
            wav_path=str(pair.wav_path),
            reference_text=pair.reference_text,
            prediction=prediction,
            wer=_wer(pair.reference_text, prediction) if not error else 1.0,
            cer=_cer(pair.reference_text, prediction) if not error else 1.0,
            execution_time_s=elapsed,
            prompt_tokens=prompt_tokens,
            output_tokens=output_tokens,
            total_tokens=total_tokens,
            error=error,
        )
        metrics.append(row)
        checkpoint[pair.stem] = asdict(row)
        _save_checkpoint(args.checkpoint_json, checkpoint)
        _append_or_create_xlsx(args.output_xlsx, row)

        if args.progress_every > 0 and (idx % args.progress_every == 0 or idx == len(selected)):
            wall = time.perf_counter() - started_at
            print(
                (
                    f"[progress] {idx}/{len(selected)} stem={pair.stem}"
                    f" wer={row.wer:.4f} cer={row.cer:.4f}"
                    f" tokens={row.total_tokens} time_s={row.execution_time_s:.2f}"
                    f" wall_s={wall:.1f}"
                    + (f" ERROR={error[:80]}" if error else "")
                ),
                flush=True,
            )

    summary = _write_summary(
        args.summary_json,
        data_dir,
        args.model,
        metrics,
        skipped,
    )
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Evaluate Gemini STT on a folder of matching .wav + .txt files."
    )
    parser.add_argument(
        "--data-dir",
        type=Path,
        required=True,
        help="Folder with paired .wav and .txt files (same basename).",
    )
    parser.add_argument(
        "--model",
        type=str,
        default=DEFAULT_MODEL,
        help=f"Gemini model id (default: {DEFAULT_MODEL}).",
    )
    parser.add_argument(
        "--sample-size",
        type=int,
        default=None,
        help="Random sample size from all pairs (reproducible with --seed).",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=42,
        help="RNG seed for --sample-size.",
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        default=None,
        help="Process at most N pairs after sampling.",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=Path("report_gemini_stt_rows.xlsx"),
        help="Per-row metrics spreadsheet.",
    )
    parser.add_argument(
        "--summary-json",
        type=Path,
        default=Path("report_gemini_stt_summary.json"),
        help="Aggregate metrics JSON.",
    )
    parser.add_argument(
        "--checkpoint-json",
        type=Path,
        default=Path("report_gemini_stt_checkpoint.json"),
        help="Resume checkpoint (one object per stem).",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Skip stems already present in checkpoint JSON.",
    )
    parser.add_argument(
        "--delay-s",
        type=float,
        default=6.5,
        help="Seconds between API calls (free tier ~10 RPM -> 6+ seconds).",
    )
    parser.add_argument(
        "--temperature",
        type=float,
        default=0.0,
        help="Generation temperature (0 = most deterministic).",
    )
    parser.add_argument(
        "--progress-every",
        type=int,
        default=1,
        help="Print progress every N rows.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Estimate token/request usage without calling the API.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    summary = run_evaluation(args)
    if not args.dry_run:
        print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
